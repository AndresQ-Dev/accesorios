from __future__ import annotations

import hashlib
import io
import json
import re
import zipfile
from datetime import UTC, datetime, timedelta
from typing import Any
from uuid import uuid4

from flask import current_app, request
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from sqlalchemy import Connection, text

from app.backups import create_backup
from app.catalog import _audit, catalog_version, register_verified_itf_alias
from app.db import read_connection, write_connection
from app.errors import ApiError
from app.search import normalize_search_text

HEADERS = ["Código", "C.Barras", "Articulo", "Stock fisico", "Precio"]
STOCK_HEADERS = {"Stock fisico", "Stock físico"}
LIMITS = {
    "entries": 64,
    "expandedBytes": 8 * 1024 * 1024,
    "sheets": 1,
    "rows": 10_000,
    "columns": 5,
    "cellLength": 512,
}


def _invalid(message: str) -> ApiError:
    return ApiError(422, "INVALID_XLSX", message)


def _timestamp(value: datetime) -> str:
    return value.isoformat(timespec="milliseconds").replace("+00:00", "Z")


def _preflight_archive(buffer: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(buffer)) as archive:
            entries = archive.infolist()
            if len(entries) > LIMITS["entries"]:
                raise _invalid("XLSX archive contains too many entries.")
            expanded = 0
            for entry in entries:
                expanded += entry.file_size
                if expanded > LIMITS["expandedBytes"] or re.search(
                    r"(^|/)vbaProject\.bin$", entry.filename, re.IGNORECASE
                ):
                    raise _invalid("XLSX archive exceeds safe limits or contains macros.")
                if entry.flag_bits & 0x1:
                    raise _invalid("Encrypted XLSX archives are not allowed.")
    except zipfile.BadZipFile:
        raise _invalid("Upload is not a valid ZIP archive.") from None


def read_xlsx_upload() -> bytes:
    content_type = request.headers.get("Content-Type", "")
    approved = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if approved not in content_type:
        raise ApiError(415, "UNSUPPORTED_MEDIA_TYPE", "Requests must use the XLSX content type.")
    declared = request.content_length or 0
    limit = current_app.config["XLSX_UPLOAD_LIMIT"]
    if declared > limit:
        raise ApiError(413, "XLSX_TOO_LARGE", "XLSX upload exceeds the configured limit.")
    body = request.get_data(cache=False)
    if len(body) > limit:
        raise ApiError(413, "XLSX_TOO_LARGE", "XLSX upload exceeds the configured limit.")
    if not body:
        raise _invalid("XLSX upload is required.")
    return body


def _text(cell: Cell | MergedCell, name: str, *, required: bool = True) -> str:
    if cell.data_type == "f":
        raise _invalid("Formulas are not allowed in XLSX imports.")
    value = cell.value
    if value is not None and not isinstance(value, (str, int, float)):
        raise _invalid(f"{name} must be text or a number.")
    if isinstance(value, bool):
        raise _invalid(f"{name} must be text or a number.")
    result = " ".join(str(value if value is not None else "").strip().split())
    if len(result) > LIMITS["cellLength"]:
        raise _invalid(f"{name} exceeds the configured cell length.")
    if required and not result:
        raise _invalid(f"{name} is required.")
    return result


def _whole(cell: Cell | MergedCell, name: str, *, required: bool = True) -> int | None:
    value = _text(cell, name, required=required)
    if not value and not required:
        return None
    if not re.fullmatch(r"\d+", value):
        raise _invalid(f"{name} must be a non-negative whole number.")
    number = int(value)
    if number > 9_007_199_254_740_991:
        raise _invalid(f"{name} is outside the safe integer range.")
    return number


def _barcode(cell: Cell | MergedCell) -> str | None:
    value = _text(cell, "C.Barras", required=False)
    if not value:
        return None
    scientific_format = re.search(r"[0#?][Ee][+-]?[0#?]", str(cell.number_format))
    if re.search(r"e[+-]?\d+$", value, re.IGNORECASE) or scientific_format:
        raise _invalid("C.Barras must not use scientific notation.")
    return value


def _blank_row(cells: list[Cell | MergedCell]) -> bool:
    return all(_text(cell, "Import column", required=False) == "" for cell in cells)


def _parse_workbook(buffer: bytes) -> list[dict[str, Any]]:
    _preflight_archive(buffer)
    try:
        workbook = load_workbook(io.BytesIO(buffer), read_only=False, data_only=False)
    except Exception:
        raise _invalid("XLSX workbook could not be parsed.") from None
    try:
        if len(workbook.worksheets) != LIMITS["sheets"]:
            raise _invalid("XLSX must contain exactly one worksheet.")
        sheet = workbook.worksheets[0]
        if sheet.max_row < 2 or sheet.max_row - 1 > LIMITS["rows"] or sheet.max_column > LIMITS["columns"]:
            raise _invalid("XLSX exceeds worksheet limits.")
        received = [_text(sheet.cell(1, index), f"Header {index}") for index in range(1, 6)]
        if received[:3] != HEADERS[:3] or received[3] not in STOCK_HEADERS or received[4] != HEADERS[4]:
            raise _invalid("XLSX headers must exactly match the approved worksheet columns.")
        rows: list[dict[str, Any]] = []
        codes: set[str] = set()
        barcodes: set[str] = set()
        for index in range(2, sheet.max_row + 1):
            if _blank_row([sheet.cell(index, column) for column in range(1, 6)]):
                continue
            row = {
                "code": _text(sheet.cell(index, 1), "Código"),
                "barcode": _barcode(sheet.cell(index, 2)),
                "article": _text(sheet.cell(index, 3), "Articulo"),
                "stock": _whole(sheet.cell(index, 4), "Stock fisico", required=False),
                "priceArs": _whole(sheet.cell(index, 5), "Precio", required=False),
            }
            code_key = normalize_search_text(str(row["code"]))
            barcode_key = normalize_search_text(str(row["barcode"])) if row["barcode"] else None
            if code_key in codes or (barcode_key is not None and barcode_key in barcodes):
                raise _invalid("XLSX contains duplicate normalized codes or barcodes.")
            codes.add(code_key)
            if barcode_key is not None:
                barcodes.add(barcode_key)
            rows.append(row)
        return rows
    finally:
        workbook.close()


def preview_xlsx(buffer: bytes, actor_hash: str) -> dict[str, Any]:
    rows = _parse_workbook(buffer)
    with write_connection() as connection:
        connection.execute(
            text("DELETE FROM import_previews WHERE expires_at <= :now"),
            {"now": _timestamp(datetime.now(UTC))},
        )
        existing = _existing_products_by_code(connection)
        creates = 0
        updates = 0
        for row in rows:
            product = existing.get(normalize_search_text(row["code"]))
            if product is None:
                creates += 1
            elif _product_has_import_changes(product, row):
                updates += 1
        expires_at = datetime.now(UTC) + timedelta(seconds=current_app.config["PREVIEW_SECONDS"])
        reference = str(uuid4())
        content_hash = hashlib.sha256(buffer).hexdigest()
        base_version = catalog_version(connection)
        connection.execute(
            text(
                "INSERT INTO import_previews "
                "(reference, actor_session_hash, content_hash, base_catalog_version, expires_at, rows_json) "
                "VALUES (:reference, :actor, :hash, :version, :expires, :rows)"
            ),
            {
                "reference": reference,
                "actor": actor_hash,
                "hash": content_hash,
                "version": base_version,
                "expires": _timestamp(expires_at),
                "rows": json.dumps(rows, ensure_ascii=False, separators=(",", ":")),
            },
        )
    return {
        "previewReference": reference,
        "contentHash": content_hash,
        "baseCatalogVersion": base_version,
        "diff": {"creates": creates, "updates": updates},
        "expiresAt": _timestamp(expires_at),
        "rows": rows,
    }


def _load_preview(reference: str, actor_hash: str) -> dict[str, Any]:
    with read_connection() as connection:
        row = connection.execute(
            text(
                "SELECT content_hash, base_catalog_version, expires_at, rows_json "
                "FROM import_previews WHERE reference = :reference AND actor_session_hash = :actor"
            ),
            {"reference": reference, "actor": actor_hash},
        ).mappings().one_or_none()
    if row is None:
        raise ApiError(409, "PREVIEW_NOT_FOUND", "The import preview is no longer available.")
    expires_at = datetime.fromisoformat(row["expires_at"].replace("Z", "+00:00"))
    if expires_at <= datetime.now(UTC):
        with write_connection() as connection:
            connection.execute(
                text("DELETE FROM import_previews WHERE reference = :reference"),
                {"reference": reference},
            )
        raise ApiError(409, "PREVIEW_EXPIRED", "The import preview has expired.")
    return {
        "reference": reference,
        "contentHash": row["content_hash"],
        "baseCatalogVersion": row["base_catalog_version"],
        "rows": json.loads(row["rows_json"]),
    }


def _parse_confirmation(payload: Any) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise ApiError(422, "INVALID_CONFIRMATION", "Import confirmation is invalid.")
    if set(payload) != {"previewReference", "contentHash", "baseCatalogVersion"}:
        raise ApiError(422, "INVALID_CONFIRMATION", "Import confirmation is invalid.")
    reference = payload["previewReference"]
    content_hash = payload["contentHash"]
    version = payload["baseCatalogVersion"]
    if (
        not isinstance(reference, str)
        or not isinstance(content_hash, str)
        or not re.fullmatch(r"[a-f0-9]{64}", content_hash)
        or isinstance(version, bool)
        or not isinstance(version, int)
    ):
        raise ApiError(422, "INVALID_CONFIRMATION", "Import confirmation is invalid.")
    return payload


def _existing_products_by_code(connection: Connection) -> dict[str, dict[str, Any]]:
    return {
        row["codeKey"]: dict(row)
        for row in connection.execute(
            text(
                "SELECT id, code, code_key AS codeKey, barcode, article, stock, "
                "price_ars AS priceArs FROM products"
            )
        ).mappings()
    }


def _product_has_import_changes(product: dict[str, Any], row: dict[str, Any]) -> bool:
    return any((
        product["code"] != row["code"],
        product["barcode"] != row["barcode"],
        product["article"] != row["article"],
        product["stock"] != row["stock"],
        product["priceArs"] != row["priceArs"],
    ))


def _assert_barcode_collisions(connection: Connection, rows: list[dict[str, Any]]) -> None:
    products = [dict(row) for row in connection.execute(
        text("SELECT id, code_key AS codeKey, barcode FROM products")
    ).mappings()]
    aliases = list(connection.execute(text("SELECT alias FROM barcode_aliases")).scalars())
    for row in rows:
        if not row["barcode"]:
            continue
        code_key = normalize_search_text(row["code"])
        barcode_key = normalize_search_text(row["barcode"])
        canonical = next(
            (
                product for product in products
                if product["barcode"] is not None and normalize_search_text(product["barcode"]) == barcode_key
            ),
            None,
        )
        if canonical and canonical["codeKey"] != code_key:
            raise ApiError(409, "BARCODE_COLLISION", "An imported barcode belongs to another product.")
        if any(normalize_search_text(alias) == barcode_key for alias in aliases):
            raise ApiError(409, "BARCODE_COLLISION", "An imported barcode collides with a registered alias.")


def confirm_xlsx(payload: Any, actor_hash: str) -> dict[str, int]:
    confirmation = _parse_confirmation(payload)
    preview = _load_preview(confirmation["previewReference"], actor_hash)
    if (
        confirmation["contentHash"] != preview["contentHash"]
        or confirmation["baseCatalogVersion"] != preview["baseCatalogVersion"]
    ):
        raise ApiError(409, "PREVIEW_MISMATCH", "The import confirmation does not match its preview.")
    with read_connection() as connection:
        if catalog_version(connection) != preview["baseCatalogVersion"]:
            raise ApiError(409, "REVISION_CONFLICT", "The catalog changed after preview.")
    create_backup()
    with write_connection() as connection:
        if catalog_version(connection) != preview["baseCatalogVersion"]:
            raise ApiError(409, "REVISION_CONFLICT", "The catalog changed after preview.")
        _assert_barcode_collisions(connection, preview["rows"])
        creates = 0
        updates = 0
        existing = _existing_products_by_code(connection)
        for row in preview["rows"]:
            code_key = normalize_search_text(row["code"])
            product = existing.get(code_key)
            if product is not None:
                if _product_has_import_changes(product, row):
                    connection.execute(
                        text(
                            "UPDATE products SET code = :code, barcode = :barcode, article = :article, "
                            "article_key = :article_key, stock = :stock, price_ars = :price, "
                            "revision = revision + 1, updated_at = CURRENT_TIMESTAMP WHERE id = :id"
                        ),
                        {
                            "code": row["code"], "barcode": row["barcode"], "article": row["article"],
                            "article_key": normalize_search_text(row["article"]), "stock": row["stock"],
                            "price": row["priceArs"], "id": product["id"],
                        },
                    )
                    updates += 1
            else:
                connection.execute(
                    text(
                        "INSERT INTO products "
                        "(code, code_key, barcode, brand, brand_key, article, article_key, "
                        "category_id, stock, price_ars) VALUES (:code, :key, :barcode, NULL, NULL, "
                        ":article, :article_key, NULL, :stock, :price)"
                    ),
                    {
                        "code": row["code"], "key": code_key, "barcode": row["barcode"],
                        "article": row["article"], "article_key": normalize_search_text(row["article"]),
                        "stock": row["stock"], "price": row["priceArs"],
                    },
                )
                creates += 1
            if row["barcode"]:
                try:
                    register_verified_itf_alias(connection, row["barcode"])
                except ValueError:
                    raise ApiError(
                        409, "BARCODE_COLLISION", "The verified barcode alias belongs to another product."
                    ) from None
        new_version = (
            preview["baseCatalogVersion"] + 1 if creates or updates else preview["baseCatalogVersion"]
        )
        connection.execute(
            text("UPDATE catalog_metadata SET catalog_version = :version WHERE id = 1"),
            {"version": new_version},
        )
        connection.execute(
            text(
                "INSERT INTO import_runs "
                "(actor_session_hash, content_hash, base_catalog_version, catalog_version, row_count) "
                "VALUES (:actor, :hash, :base, :version, :count)"
            ),
            {
                "actor": actor_hash, "hash": preview["contentHash"],
                "base": preview["baseCatalogVersion"], "version": new_version,
                "count": len(preview["rows"]),
            },
        )
        _audit(
            connection, actor_hash, "import.confirmed",
            {"creates": creates, "updates": updates, "catalogVersion": new_version},
        )
        connection.execute(
            text("DELETE FROM import_previews WHERE reference = :reference"),
            {"reference": preview["reference"]},
        )
    return {"catalogVersion": new_version, "creates": creates, "updates": updates}
